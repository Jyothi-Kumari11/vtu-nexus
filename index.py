# ================================================================
#   VTU RESULTS SCRAPER - Standalone, Self-contained
# ================================================================
import os, re, sys, time, logging, shutil
import pytesseract, cv2
import numpy as np
import pandas as pd
from collections import Counter
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.chart import BarChart, PieChart, Reference

# ── CONFIG ────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
log = logging.getLogger("vtu")

_tess = shutil.which("tesseract")
if _tess:
    pytesseract.pytesseract.tesseract_cmd = _tess
else:
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

VTU_URL        = "https://results.vtu.ac.in/D25J26Ecbcs/index.php"
INPUT_FILE     = "usn_list.xlsx"
RAW_OUTPUT     = "results_output.xlsx"
FORMATTED_OUT  = "formatted_result_sheet.xlsx"
FINAL_OUTPUT   = "final_report.xlsx"
CAPTCHA_IMG    = "captcha.png"
SEM_DIGIT      = '3'          # Change for other semesters

# ── CAPTCHA SOLVER ────────────────────────────────────────────
_OCR_CFG = r"--oem 3 --psm 8 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789"

def _clean(t):
    return re.sub(r'[^A-Za-z0-9]', '', t).strip()

def solve_captcha(path):
    img = cv2.imread(path)
    if img is None:
        return ""
    g = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    candidates = []
    # 4 preprocessing strategies
    for proc in [
        lambda g: cv2.threshold(g, 127, 255, cv2.THRESH_BINARY)[1],
        lambda g: cv2.threshold(g, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1],
        lambda g: cv2.adaptiveThreshold(cv2.fastNlMeansDenoising(
            cv2.resize(g, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC), h=15),
            255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2),
        lambda g: cv2.threshold(cv2.medianBlur(g, 3), 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1],
    ]:
        try:
            t = _clean(pytesseract.image_to_string(proc(g), config=_OCR_CFG))
            if 3 <= len(t) <= 8:
                candidates.append(t)
        except Exception:
            pass
    if not candidates:
        return _clean(pytesseract.image_to_string(g, config=_OCR_CFG))
    return Counter(candidates).most_common(1)[0][0]

# ── BROWSER ───────────────────────────────────────────────────
def start_browser():
    opts = webdriver.ChromeOptions()
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    # Container/Linux headless mode
    if os.name != "nt":
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
        opts.add_argument("--disable-gpu")
    driver = webdriver.Chrome(options=opts)
    if os.name == "nt":
        driver.maximize_window()
    log.info("Browser started.")
    return driver

# ── STEP 1: READ USN LIST ─────────────────────────────────────
log.info(f"Reading USN list from {INPUT_FILE}")
df_usn = pd.read_excel(INPUT_FILE)
usn_list = df_usn.iloc[:, 0].dropna().astype(str).str.strip().tolist()
log.info(f"Found {len(usn_list)} USNs")

# ── STEP 2: LOAD CACHE ────────────────────────────────────────
all_results = []
existing_usns = set()
if os.path.exists(RAW_OUTPUT):
    try:
        old = pd.read_excel(RAW_OUTPUT)
        all_results = old.to_dict('records')
        existing_usns = set(old["USN"].astype(str).str.strip())
        log.info(f"Cache loaded: {len(existing_usns)} already done, skipping them.")
    except Exception as e:
        log.warning(f"Cache load failed: {e}")

# Check report-only flag
report_only = "--report-only" in sys.argv

if report_only:
    log.info("Report-only mode active. Skipping scraping step.")
    if not all_results:
        log.error("No results cache found to build report from. Exiting.")
        sys.exit(1)
else:
    # ── STEP 3: SCRAPE ────────────────────────────────────────────
    driver = start_browser()

    SUBJ_PATTERN = re.compile(
        r'([A-Z]{2,6}\d{3,4}[A-Z]?)\s+(.+?)\s+'
        r'(\d{1,3}|AB|ABS)\s+(\d{1,3}|AB|ABS)\s+(\d{1,3}|AB|ABS)\s+'
        r'([A-Z]{1,4})\s+(\d{4}-\d{2}-\d{2})'
    )

    for idx, usn in enumerate(usn_list, 1):
        if usn in existing_usns:
            log.info(f"[{idx}/{len(usn_list)}] {usn} — SKIPPED (cached)")
            continue

        log.info(f"[{idx}/{len(usn_list)}] Processing: {usn}")
        retry = 0

        while True:
            try:
                driver.get(VTU_URL)
                usn_box = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.NAME, "lns")))
                usn_box.clear()
                usn_box.send_keys(usn)

                # CAPTCHA
                cap_img = driver.find_element(By.XPATH, "//img[contains(@src,'captcha')]")
                cap_img.screenshot(CAPTCHA_IMG)
                captcha_text = solve_captcha(CAPTCHA_IMG)
                log.info(f"  CAPTCHA guess: {captcha_text}")

                cap_box = driver.find_element(By.NAME, "captchacode")
                cap_box.clear()
                cap_box.send_keys(captcha_text)
                driver.find_element(By.XPATH, "//input[@type='submit']").click()

                # Alert check
                try:
                    alert = WebDriverWait(driver, 3).until(EC.alert_is_present())
                    msg = alert.text.lower()
                    alert.accept()
                    if "captcha" in msg:
                        log.warning("  Wrong CAPTCHA, retrying...")
                        retry += 1
                        time.sleep(1)
                        continue
                    else:
                        log.warning(f"  USN {usn} invalid/no data. Skipping.")
                        break
                except TimeoutException:
                    pass

                # Verify results page
                body = driver.find_element(By.TAG_NAME, "body").text
                if "Semester" not in body:
                    log.warning("  Results page not loaded, retrying...")
                    retry += 1
                    time.sleep(2)
                    continue

                # Extract name & semester
                name, sem = "Unknown", "Unknown"
                for line in body.split("\n"):
                    if "Student Name" in line: name = line.split(":")[-1].strip()
                    if "Semester" in line:     sem  = line.split(":")[-1].strip()

                # Extract subjects
                seen = set()
                added = 0
                for m in SUBJ_PATTERN.findall(body):
                    code = m[0]
                    if code in seen:
                        continue
                    # Filter by semester digit
                    d = re.search(r'\d', code)
                    if d and d.group() != SEM_DIGIT:
                        continue
                    seen.add(code)
                    all_results.append({
                        "USN": usn, "Name": name, "Semester": sem,
                        "SUBJECT CODE": code,
                        "SUBJECT NAME": " ".join(m[1].split()),
                        "INTERNAL MARKS": m[2], "EXTERNAL MARKS": m[3],
                        "TOTAL": m[4], "RESULT": m[5], "DATE": m[6]
                    })
                    added += 1

                log.info(f"  Extracted {added} subjects for {name}")
                break  # success

            except Exception as e:
                retry += 1
                err = str(e)[:120]
                log.warning(f"  Attempt {retry} failed: {err}")
                if any(x in err.lower() for x in ["disconnected", "session", "refused", "connection"]):
                    log.info("  Restarting browser...")
                    try: driver.quit()
                    except: pass
                    time.sleep(3)
                    driver = start_browser()
                else:
                    time.sleep(2)

    try:
        driver.quit()
    except: pass
    log.info("Browser closed.")

# ── STEP 4: SAVE RAW OUTPUT ───────────────────────────────────
raw_df = pd.DataFrame(all_results).dropna(how='all')
if not report_only:
    raw_df.to_excel(RAW_OUTPUT, index=False)
    log.info(f"Raw results saved → {RAW_OUTPUT}  ({len(raw_df)} rows)")

if raw_df.empty:
    log.error("No results extracted. Exiting.")
    sys.exit(1)

raw_df = raw_df.fillna("")

# ── STEP 5: FORMAT SUMMARY SHEET ─────────────────────────────
subject_map = {}
if 'SUBJECT CODE' in raw_df.columns and 'SUBJECT NAME' in raw_df.columns:
    subject_map = dict(zip(raw_df['SUBJECT CODE'], raw_df['SUBJECT NAME']))

unique_subjects = []
for _, row in raw_df.iterrows():
    if row["SUBJECT CODE"] and row["SUBJECT CODE"] not in unique_subjects:
        unique_subjects.append(row["SUBJECT CODE"])

h0 = ["Student Info", "Student Info"]
h1 = ["USN", "NAME"]
for code in unique_subjects:
    h0 += [code]*4
    h1 += ["Int", "Ext", "Tot", "Res"]
mi = pd.MultiIndex.from_arrays([h0, h1])

# Build student map
smap = {}
for u in usn_list:
    u = str(u).strip()
    d = {("Student Info","USN"): u, ("Student Info","NAME"): ""}
    for c in unique_subjects:
        d[(c,"Int")] = d[(c,"Ext")] = d[(c,"Tot")] = d[(c,"Res")] = ""
    smap[u] = d

for _, row in raw_df.iterrows():
    u = str(row["USN"]).strip()
    if u not in smap:
        smap[u] = {("Student Info","USN"): u, ("Student Info","NAME"): ""}
        for c in unique_subjects:
            smap[u][(c,"Int")] = smap[u][(c,"Ext")] = smap[u][(c,"Tot")] = smap[u][(c,"Res")] = ""
    if row["Name"]:
        smap[u][("Student Info","NAME")] = row["Name"]
    c = row["SUBJECT CODE"]
    if c:
        smap[u][(c,"Int")] = row["INTERNAL MARKS"]
        smap[u][(c,"Ext")] = row["EXTERNAL MARKS"]
        smap[u][(c,"Tot")] = row["TOTAL"]
        smap[u][(c,"Res")] = row["RESULT"]

seen_usns = set()
rows = []
for u in [str(x).strip() for x in usn_list]:
    if u in smap and u not in seen_usns:
        rows.append(smap[u]); seen_usns.add(u)
for u, d in smap.items():
    if u not in seen_usns:
        rows.append(d)

fmt_df = pd.DataFrame(rows, columns=mi)

# ── STEP 6: CALCULATE STATS ───────────────────────────────────
fmt_df.insert(0, ("Student Info","Sl. No"), range(1, len(fmt_df)+1))
tot_cols = fmt_df.xs("Tot", level=1, axis=1, drop_level=False)
fmt_df[("Performance","Total Marks")] = tot_cols.apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)

def pct(row):
    n = sum(1 for c in unique_subjects if str(row[(c,"Tot")]).strip())
    return round(row[("Performance","Total Marks")]/n, 2) if n else 0.0
fmt_df[("Performance","Percentage %")] = fmt_df.apply(pct, axis=1)

def stats(row):
    p = sum(1 for c in unique_subjects if str(row[(c,"Res")]).strip().upper()=="P")
    f = sum(1 for c in unique_subjects if str(row[(c,"Res")]).strip().upper() in {"F","AB","ABS","FAIL"})
    return p, f
st = fmt_df.apply(stats, axis=1)
fmt_df[("Performance","Passed")] = [s[0] for s in st]
fmt_df[("Performance","Failed")] = [s[1] for s in st]
fmt_df[("Performance","Class")] = fmt_df.apply(
    lambda r: "FAIL" if r[("Performance","Failed")]>0 else ("FCD" if r[("Performance","Percentage %")]>80 else "PASS"), axis=1)

# ── STEP 7: ANALYSIS ──────────────────────────────────────────
analysis_rows = []
for i, subj in enumerate(unique_subjects, 1):
    passed = (fmt_df[(subj,"Res")]=="P").sum()
    total  = len(fmt_df)
    analysis_rows.append([i, subj, subject_map.get(subj,""), total, passed,
                          round(passed/total*100,2) if total else 0,
                          round((total-passed)/total*100,2) if total else 0])
analysis_df = pd.DataFrame(analysis_rows,
    columns=["Sl No","Subject Code","Subject Name","Appeared","Passed","Pass %","Fail %"])

top3 = fmt_df.sort_values([("Performance","Total Marks"),("Performance","Percentage %")], ascending=False).head(3)
top3_list = [{"Rank":i+1,"USN":row[("Student Info","USN")],"NAME":row[("Student Info","NAME")],
              "Total":row[("Performance","Total Marks")],"Pct":row[("Performance","Percentage %")]}
             for i, (_, row) in enumerate(top3.iterrows())]
top3_df = pd.DataFrame(top3_list)

total_s = len(fmt_df)
t_pass  = (fmt_df[("Performance","Class")]!="FAIL").sum()
t_fail  = (fmt_df[("Performance","Class")]=="FAIL").sum()
pass_pct = round(t_pass/total_s*100, 2) if total_s else 0
abstract_df = pd.DataFrame({
    "Metric": ["Total Students","Total Pass","Total Fail","Pass %"],
    "Value":  [total_s, t_pass, t_fail, f"{pass_pct}%"]
})

fail_mask = fmt_df[("Performance","Class")]=="FAIL"
fail_data = []
for _, row in fmt_df[fail_mask].iterrows():
    fails = [f"{c} ({subject_map.get(c,'')})" for c in unique_subjects
             if str(row[(c,"Res")]).strip().upper() in {"F","AB","ABS","FAIL"}]
    fail_data.append({"USN":row[("Student Info","USN")],
                      "NAME":row[("Student Info","NAME")],
                      "Failed Subjects":", ".join(fails)})
fail_df = pd.DataFrame(fail_data) if fail_data else pd.DataFrame(columns=["USN","NAME","Failed Subjects"])

# Summary row
sr = {col:"" for col in fmt_df.columns}
sr[("Student Info","Sl. No")] = "SUMMARY"
sr[("Student Info","USN")]    = f"PASS:{t_pass}"
sr[("Student Info","NAME")]   = f"FAIL:{t_fail}"
sr[("Performance","Class")]   = f"{pass_pct}%"
fmt_df = pd.concat([fmt_df, pd.DataFrame([sr])], ignore_index=True)

# Save formatted sheet
flat = fmt_df.copy()
flat.columns = [f"{a}_{b}" if a not in ("Student Info","Performance") else b for a,b in flat.columns]
flat.to_excel(FORMATTED_OUT, index=False)
log.info(f"Formatted sheet saved → {FORMATTED_OUT}")

# ── STEP 8: FINAL REPORT ──────────────────────────────────────
with pd.ExcelWriter(FINAL_OUTPUT, engine="openpyxl") as writer:
    res_flat = fmt_df.copy()
    res_flat.columns = [b for a,b in res_flat.columns]
    res_flat.to_excel(writer, sheet_name="Results", index=False)

    pd.DataFrame([["SUBJECT-WISE ANALYSIS"]]).to_excel(writer, sheet_name="Analysis", startrow=0, index=False, header=False)
    analysis_df.to_excel(writer, sheet_name="Analysis", startrow=1, index=False)
    cur = len(analysis_df)+4
    pd.DataFrame([["TOP 3 SCORERS"]]).to_excel(writer, sheet_name="Analysis", startrow=cur, index=False, header=False)
    top3_df.to_excel(writer, sheet_name="Analysis", startrow=cur+1, index=False)
    cur += len(top3_df)+4
    pd.DataFrame([["PERFORMANCE ABSTRACT"]]).to_excel(writer, sheet_name="Analysis", startrow=cur, index=False, header=False)
    abstract_df.to_excel(writer, sheet_name="Analysis", startrow=cur+1, index=False)
    cur += len(abstract_df)+4
    pd.DataFrame([["FAILURE DETAILS"]]).to_excel(writer, sheet_name="Analysis", startrow=cur, index=False, header=False)
    fail_df.to_excel(writer, sheet_name="Analysis", startrow=cur+1, index=False)
    abs_data_start = len(analysis_df)+4+len(top3_df)+4+6  # for chart reference

# ── STEP 9: OPENPYXL FORMATTING ──────────────────────────────
wb = load_workbook(FINAL_OUTPUT)
ws = wb["Results"]

thin = Border(*[Side(style="thin")]*0,
              left=Side(style="thin"), right=Side(style="thin"),
              top=Side(style="thin"), bottom=Side(style="thin"))
bold = Font(bold=True)
hdr_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Insert 3 header rows
ws.insert_rows(1, 3)
title = f"VTU Result Sheet — Semester {SEM_DIGIT} — Academic Year: 2024-25"
ws.cell(1,1).value = title
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ws.max_column)
ws.cell(1,1).font = Font(bold=True, size=13)

static = ["Sl. No","USN","NAME"]
for i, lbl in enumerate(static, 1):
    ws.cell(2,i).value = lbl
    ws.merge_cells(start_row=2, start_column=i, end_row=4, end_column=i)

col = 4
for code in unique_subjects:
    ws.cell(2,col).value = code
    ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+3)
    nm = subject_map.get(code,"")
    abbr = "".join(w[0] for w in nm.split() if w and w[0].isupper()) or code[:5]
    ws.cell(3,col).value = abbr
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+3)
    for i, lbl in enumerate(["Int","Ext","Total","Res"]):
        ws.cell(4, col+i).value = lbl
    col += 4

for i, lbl in enumerate(["Total Marks","Percentage %","Passed","Failed","Class"]):
    ws.cell(2, col+i).value = lbl
    ws.merge_cells(start_row=2, start_column=col+i, end_row=4, end_column=col+i)

for row in ws.iter_rows():
    for cell in row:
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if cell.row == 1:
            cell.font = Font(bold=True, size=13)
        elif cell.row <= 4:
            cell.font = bold; cell.fill = hdr_fill

# Colour P/F cells
for r in range(5, ws.max_row+1):
    for c in range(4, ws.max_column+1):
        hv = str(ws.cell(4,c).value)
        h2 = str(ws.cell(2,c).value)
        cell = ws.cell(r,c)
        v = str(cell.value).strip().upper()
        if hv == "Res":
            if v == "P":   cell.fill = green_fill; cell.font = bold
            elif v in {"F","AB","ABS","FAIL"}: cell.fill = red_fill; cell.font = bold
        if h2 == "Class":
            if v in {"PASS","FCD"}: cell.fill = green_fill; cell.font = bold
            elif v == "FAIL":       cell.fill = red_fill; cell.font = bold

# Analysis sheet
ws2 = wb["Analysis"]
ws2.insert_rows(1, 4)
ws2.cell(1,1).value = "VTU RESULTS ANALYSIS REPORT"
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
ws2.cell(1,1).font = Font(bold=True, size=18, color="1F497D")
ws2.cell(1,1).alignment = Alignment(horizontal="center", vertical="center")
ws2.cell(3,1).value = f"Semester: {SEM_DIGIT}  |  Academic Year: 2024-25  |  Overall Pass Rate: {pass_pct}%"
ws2.merge_cells(start_row=3, start_column=1, end_row=3, end_column=8)
ws2.cell(3,1).font = Font(bold=True, size=11, color="475569")
ws2.cell(3,1).alignment = Alignment(horizontal="center", vertical="center")

# 1. Page orientation and margins for strict A4 Portrait fit
ws2.page_setup.orientation = ws2.ORIENTATION_PORTRAIT
ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
ws2.page_setup.fitToPage = True
ws2.page_setup.fitToWidth = 1
ws2.page_setup.fitToHeight = 1 # Force onto a single A4 page!

ws2.page_margins.left = 0.25
ws2.page_margins.right = 0.25
ws2.page_margins.top = 0.4
ws2.page_margins.bottom = 0.4

# 2. Explicit narrow columns to prevent squeezing and keep it clean
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 32 # subject & student names wrap beautifully
ws2.column_dimensions['D'].width = 11
ws2.column_dimensions['E'].width = 11
ws2.column_dimensions['F'].width = 10
ws2.column_dimensions['G'].width = 10

# Results sheet scaling (A4 Landscape)
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0 # standard vertical flow

# Auto-fit only Results sheet columns
from openpyxl.utils import get_column_letter
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row <= 4:
            continue
        val = str(cell.value or '')
        if len(val) > max_len:
            max_len = len(val)
    ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

# Apply beautiful styles and borders to Analysis tables dynamically
sec_fill = PatternFill(start_color="365F91", end_color="365F91", fill_type="solid")
white_bold = Font(color="FFFFFF", bold=True, size=10)
section_font = Font(bold=True, size=12, color="1F497D")

# Find dynamic header rows in Analysis sheet
toppers_hdr_row = None
abstract_hdr_row = None
failures_hdr_row = None

for r in range(5, ws2.max_row+1):
    val = str(ws2.cell(r, 1).value).strip()
    if "TOP 3 SCORERS" in val:
        toppers_hdr_row = r
    elif "PERFORMANCE ABSTRACT" in val:
        abstract_hdr_row = r
    elif "FAILURE DETAILS" in val:
        failures_hdr_row = r

# Loop through all cells in Analysis sheet and format them beautifully!
for r in range(5, ws2.max_row+1):
    val_a = str(ws2.cell(r, 1).value or '').strip()
    is_section_title = val_a in ["SUBJECT-WISE ANALYSIS", "TOP 3 SCORERS", "PERFORMANCE ABSTRACT", "FAILURE DETAILS"]
    
    for c in range(1, 8):
        cell = ws2.cell(r, c)
        
        if is_section_title:
            if c == 1:
                cell.font = section_font
                cell.alignment = Alignment(vertical="center", horizontal="left")
            continue
            
        if cell.value is None:
            continue
            
        is_hdr = (r == 6) or \
                 (toppers_hdr_row and r == toppers_hdr_row + 1) or \
                 (abstract_hdr_row and r == abstract_hdr_row + 1) or \
                 (failures_hdr_row and r == failures_hdr_row + 1)
                 
        if is_hdr:
            cell.fill = sec_fill
            cell.font = white_bold
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        else:
            cell.border = thin
            align_h = "left" if c == 3 else "center"
            cell.alignment = Alignment(vertical="center", horizontal=align_h, wrap_text=True)

# ── ADD CHARTS & SIGNATURES Underneath for perfect A4 printing ──
chart_start_row = 15
if failures_hdr_row:
    chart_start_row = failures_hdr_row + len(fail_df) + 4
elif abstract_hdr_row:
    chart_start_row = abstract_hdr_row + 8

# 1. Subject-Wise Pass % Bar Chart
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Subject Pass Rates (%)"
chart.y_axis.title = "Pass Percentage"
chart.x_axis.title = "Subject Code"
chart.width = 16
chart.height = 9
chart.legend = None

data = Reference(ws2, min_col=6, min_row=6, max_row=6 + len(unique_subjects))
cats = Reference(ws2, min_col=2, min_row=7, max_row=6 + len(unique_subjects))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws2.add_chart(chart, f"A{chart_start_row}")

# 2. Overall Pass/Fail Pie Chart
if abstract_hdr_row:
    pie = PieChart()
    pie.title = "Overall Result Distribution"
    pie.width = 14
    pie.height = 9
    
    pie_data = Reference(ws2, min_col=2, min_row=abstract_hdr_row + 3, max_row=abstract_hdr_row + 4)
    pie_cats = Reference(ws2, min_col=1, min_row=abstract_hdr_row + 3, max_row=abstract_hdr_row + 4)
    
    pie.add_data(pie_data)
    pie.set_categories(pie_cats)
    ws2.add_chart(pie, f"E{chart_start_row}")

# 3. Dynamic Signatures placed nicely below the charts
sig_row = chart_start_row + 19
ws2.cell(sig_row, 1).value = "Signature of Class Teacher: _____________"
ws2.cell(sig_row, 1).font = Font(bold=True)
ws2.cell(sig_row, 5).value = "Signature of HOD: _____________"
ws2.cell(sig_row, 5).font = Font(bold=True)

wb.save(FINAL_OUTPUT)
log.info(f"Final report saved with A4 charts and signatures → {FINAL_OUTPUT}")

print("\n" + "="*50)
print("ALL DONE!")
print(f"  Raw data      : {RAW_OUTPUT}")
print(f"  Formatted     : {FORMATTED_OUT}")
print(f"  Final report  : {FINAL_OUTPUT}")
print(f"  Students      : {total_s} | Pass: {t_pass} | Fail: {t_fail} | Pass%: {pass_pct}%")
print("="*50)